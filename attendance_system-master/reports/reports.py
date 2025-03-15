import os
from fastapi import HTTPException
from sqlalchemy.sql import func, and_
from sqlmodel import select, Session
from datetime import date, timedelta
from openpyxl import load_workbook
from starlette.responses import StreamingResponse
from fastapi import Response
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook
from app.faculty.models import Faculty
from app.students.models import Student, Attendance

from fastapi.responses import FileResponse


def faculty_report(session: Session, faculty_id: int, start_date: date, is_male: bool = True):
    faculty = session.get(Faculty, faculty_id)
    if not faculty:
        raise HTTPException(status_code=400, detail="Invalid faculty ID")

    statement = (
        select(Student.id, Student.seq_number, Student.name, Student.notes)
        .where(Student.faculty_id == faculty_id, Student.is_male == is_male)
        .order_by(Student.seq_number)
    )
    students = session.exec(statement).all()

    dates = [(start_date + timedelta(days=i + (1 if i > 5 else 0))) for i in range(12)]

    data = []
    for i_idx, student in enumerate(students):
        student_id = student[0]
        student_seq = student[1]
        student_name = student[2]
        student_notes = student[3]

        data.append([student_seq, student_name] + [""] * 12 + [student_notes if student_notes else ""])
        for j_idx, day_date in enumerate(dates):
            statement = (
                select(Attendance.id)
                .where(
                    Attendance.faculty_id == faculty_id,
                    Attendance.student_id == student_id,
                    Attendance.leave_time.is_not(None),
                    func.date(Attendance.leave_time) == day_date
                )
            )
            result = session.exec(statement).first()
            if result:
                data[i_idx][j_idx + 2] = "✔"

    # Create an Excel file dynamically in memory
    output = BytesIO()
    wb = load_workbook("reports/templates/faculty_report.xlsx")
    ws = wb.active
    for row in data:
        ws.append(row)

    wb.save(output)
    output.seek(0)  # Move cursor to the beginning of the file
    # file_path = "sdafasdf.xlsx"
    # with open(file_path, "wb") as f:
    #     f.write(output.getbuffer())
    # Return file as response without saving
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={
                                 "Content-Disposition": f'attachment; filename="faculty_report_{faculty_id}.xlsx"'})


from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, func, and_
from datetime import date
from io import BytesIO
from openpyxl import load_workbook


def day_report(session: Session, day_date: date = date.today(), is_male: bool = True):
    statement = (
        select(Faculty.name, func.count(Attendance.id))
        .select_from(Faculty)
        .join(Attendance, Faculty.id == Attendance.faculty_id, isouter=True)  # Left Outer Join
        .where(
            and_(
                Faculty.is_male == is_male,
                func.coalesce(Attendance.is_male, is_male) == is_male,  # Handle NULL cases
                func.coalesce(func.date(Attendance.leave_time), day_date) == day_date  # Handle NULL dates
            )
        )
        .group_by(Faculty.id)
        .order_by(Faculty.id)
    )

    facultyAttendance = session.exec(statement).all()
    # Load the existing template with formatting
    template_path = "reports/templates/day_report.xlsx"
    wb = load_workbook(template_path)
    ws = wb.active

    # Append the fetched data
    for row in facultyAttendance:
        ws.append([row[0], row[1], ""])  # Faculty Name | Attendance Count | Empty Column
    # Save to memory instead of disk
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return file as response
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="day_report_{day_date}.xlsx"'}
    )


def general_report(session: Session, start_date: date, is_male: bool = True):
    faculties = session.exec(select(Faculty).where(Faculty.is_male == is_male)).all()
    dates = [start_date + timedelta(days=i + (1 if i > 5 else 0)) for i in range(12)]

    data = []
    for faculty in faculties:
        faculty_data = [faculty.name] + [0] * 12 + [""]
        for j_idx, report_date in enumerate(dates):
            attendance_count = session.exec(
                select(func.count(Attendance.id)).where(
                    Attendance.faculty_id == faculty.id,
                    Attendance.is_male == is_male,
                    func.date(Attendance.leave_time) == report_date
                )
            ).one_or_none() or 0
            faculty_data[j_idx + 1] = attendance_count
        data.append(faculty_data)

    # Load the pre-styled Excel template
    wb = load_workbook("reports/templates/general_report.xlsx")
    ws = wb.active  # Select the active worksheet

    # Find the first empty row after existing headers
    start_row = ws.max_row + 1

    # Append data to the existing sheet
    for row in data:
        ws.append(row)

    # Save the updated workbook in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return the file as a response with styling preserved
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="general_report.xlsx"'}
    )
